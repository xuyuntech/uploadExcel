# -*- coding: utf-8 -*-
import openpyxl
import codecs
import json
import requests
# import urllib
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

def readExcel(path, SheetName):
    wb = openpyxl.load_workbook(path)

    # 根据工作表名获取工作表
    sheet = wb[SheetName]

    for index in range(2, sheet.max_row):

        if sheet['A' + str(index)].value == None:
            break

        dict ={}
        dict['title'] = str(sheet['A' + str(index)].value)
        dict['description'] = str(sheet['B' + str(index)].value)
        dict['participantKey'] = str(sheet['C' + str(index)].value)
        dict['phone'] = str(sheet['D' + str(index)].value)
        dict['name'] = str(sheet['E' + str(index)].value)

        # jsonfile = codecs.open(filename="Doctor.json", mode='a', encoding='utf-8')
        # data = json.dumps(dict, ensure_ascii=False) + "\n"
        # jsonfile.write(data)

        # url = "http://localhost:3000/api/Doctor"
        # postdata = urllib.parse.urlencode(dict)
        # postdata = postdata.encode('utf-8')
        # res = urllib.request.urlopen(url,postdata)
        # print(res.status, res.reason)

        # data = json.dumps(dict, ensure_ascii=False)
        url = 'http://localhost:3000/api/Doctor'
        re = requests.post(url, dict)
        print(re.status_code)

        # print(sheet['A' + str(index)].value,"\t", end="")
        # print(sheet['B' + str(index)].value,"\t", end="")
        # print(sheet['C' + str(index)].value,"\t", end="")
        # print(sheet['D' + str(index)].value,"\t", end="")
        # print(sheet['E' + str(index)].value,"\t", end="")
        # print(sheet['F' + str(index)].value,"\t", end="")
        # print(sheet['G' + str(index)].value,"\t", end="")
        # print(sheet['H' + str(index)].value,"\t", end="")
        # print(sheet['I' + str(index)].value,"\t", end="")
        # print(sheet['J' + str(index)].value,"\t", end="")
        # print(sheet['K' + str(index)].value,"\t", end="")
        # print(sheet['L' + str(index)].value,"\t", end="")
        # print(sheet['M' + str(index)].value)

# print(data.get_sheet_names())  #输出所有的工作表名
# print(sheet.title)
# sheet02 = data.get_active_sheet()      #获取活动的工作表
# print(sheet02.title)
# print(sheet.max_row)    # 最大行数
# print(sheet.max_column) # 最大列数
# print(sheet['A3'].value)  #获取A3单元格内容
# print(sheet['A3'].column)  #获取单元格列值
# print(sheet['A3'].row)  #获取单元格行号

# def POST(data):
#     url = 'http://localhost:3030/api/Doctor'
#     requests.post(url, data)

if __name__ == '__main__':
    path = '/home/sunhui/trt-health/composer/Doctor.xlsx'  # excel表格所在路径
    SheetName = 'Sheet1'  # 工作表表名
    readExcel(path, SheetName)